"""
Day 6 Capstone — generates the raw Excel inputs, the blank starter workbook,
and the fully-solved answer key, all from one deterministic seed so every
artifact is internally consistent (row counts in the instructions doc must
match what a learner will actually see in the files).
"""
import random
from datetime import date, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

random.seed(42)

OUT_DIR = "."

# ─────────────────────────────────────────────────────────────────────────────
# TRUE (clean) source-of-truth data — the answer key is built from this,
# then messiness is layered on top to produce the learner-facing raw files.
# ─────────────────────────────────────────────────────────────────────────────

WAREHOUSES = [
    ("WH01", "Mumbai Distribution Center",   "Mumbai",    "West",  50000),
    ("WH02", "Delhi Distribution Center",    "Delhi",     "North", 42000),
    ("WH03", "Bengaluru Distribution Center","Bengaluru", "South", 38000),
    ("WH04", "Hyderabad Distribution Center","Hyderabad", "South", 30000),
    ("WH05", "Pune Distribution Center",     "Pune",      "West",  25000),
]

CATEGORIES = {
    "Electronics": ["Headphones", "Chargers", "Smart Watches"],
    "Furniture":   ["Chairs", "Tables", "Shelving"],
    "Sports":      ["Yoga Mats", "Dumbbells", "Cycling Gear"],
    "Home":        ["Cookware", "Bedding", "Lighting"],
    "Beauty":      ["Skincare", "Haircare", "Fragrance"],
}

PRODUCT_NAMES = {
    "Headphones": ["Wireless Earbuds Pro", "Over-Ear Studio Headphones", "Sport Bluetooth Buds"],
    "Chargers": ["65W USB-C Fast Charger", "Wireless Charging Pad", "Car Charger Duo"],
    "Smart Watches": ["FitTrack Watch Series 3", "FitTrack Watch Lite"],
    "Chairs": ["Ergo Mesh Office Chair", "Wooden Dining Chair Set", "Recliner Lounge Chair"],
    "Tables": ["Compact Study Desk", "Folding Dining Table", "Coffee Table Oak"],
    "Shelving": ["5-Tier Bookshelf", "Wall-Mounted Floating Shelf"],
    "Yoga Mats": ["Premium Non-Slip Yoga Mat", "Travel Yoga Mat Thin"],
    "Dumbbells": ["Adjustable Dumbbell Set 20kg", "Rubber Hex Dumbbell Pair"],
    "Cycling Gear": ["Cycling Helmet Pro", "Padded Cycling Gloves"],
    "Cookware": ["Non-Stick Frying Pan 28cm", "Stainless Steel Pot Set"],
    "Bedding": ["Cotton Bedsheet Set Queen", "Microfiber Comforter"],
    "Lighting": ["LED Desk Lamp", "Smart Bulb Pack of 4"],
    "Skincare": ["Vitamin C Face Serum", "Daily Moisturizer SPF 30"],
    "Haircare": ["Argan Oil Shampoo", "Keratin Hair Mask"],
    "Fragrance": ["Citrus Eau de Toilette", "Woody Amber Perfume"],
}

products = []
pid = 1001
for cat, subs in CATEGORIES.items():
    for sub in subs:
        for name in PRODUCT_NAMES[sub]:
            unit_cost = round(random.uniform(8.0, 220.0), 2)
            products.append((f"PRD-{pid}", name, cat, sub, unit_cost))
            pid += 1
NUM_PRODUCTS = len(products)  # 34 real products (category/subcategory list lengths vary)

START_DATE = date(2026, 6, 1)
NUM_NIGHTS = 10
dates = [START_DATE + timedelta(days=i) for i in range(NUM_NIGHTS)]

# ─────────────────────────────────────────────────────────────────────────────
# TRUE fact grain: warehouse x product x date, one row per nightly count
# ─────────────────────────────────────────────────────────────────────────────

true_fact_rows = []
for wh in WAREHOUSES:
    wh_code = wh[0]
    base_capacity_share = wh[4] / 50000.0
    for prod in products:
        pcode = prod[0]
        base_stock = int(random.uniform(80, 600) * base_capacity_share)
        stock = base_stock
        for d in dates:
            # small day-to-day drift, never below 0 in the TRUE data
            stock = max(0, stock + random.randint(-25, 20))
            batch_id = f"BATCH-{d.strftime('%Y%m%d')}"
            true_fact_rows.append({
                "warehouse_id": wh_code,
                "product_id": pcode,
                "count_date": d,
                "count_batch_id": batch_id,
                "units_on_hand": stock,
            })

true_fact_df = pd.DataFrame(true_fact_rows)
assert len(true_fact_df) == len(WAREHOUSES) * len(products) * NUM_NIGHTS  # 5*30*10 = 1500

# ─────────────────────────────────────────────────────────────────────────────
# ANSWER KEY dimension tables (clean, deduped, surrogate keys assigned)
# ─────────────────────────────────────────────────────────────────────────────

dim_warehouse = pd.DataFrame(
    [{"warehouse_sk": i + 1, "warehouse_id": w[0], "warehouse_name": w[1],
      "city": w[2], "region": w[3], "capacity_units": w[4]}
     for i, w in enumerate(WAREHOUSES)]
)

dim_product = pd.DataFrame(
    [{"product_sk": i + 1, "product_id": p[0], "product_name": p[1],
      "category": p[2], "sub_category": p[3], "unit_cost": p[4]}
     for i, p in enumerate(products)]
)

dim_date_rows = []
for d in dates:
    dim_date_rows.append({
        "date_key": int(d.strftime("%Y%m%d")),
        "date": d,
        "day_of_week": d.strftime("%A"),
        "month": d.month,
        "quarter": (d.month - 1) // 3 + 1,
        "year": d.year,
    })
dim_date = pd.DataFrame(dim_date_rows)

wh_lookup = {w[0]: i + 1 for i, w in enumerate(WAREHOUSES)}
prod_lookup = {p[0]: i + 1 for i, p in enumerate(products)}

fact_rows = []
for r in true_fact_rows:
    fact_rows.append({
        "warehouse_sk": wh_lookup[r["warehouse_id"]],
        "product_sk": prod_lookup[r["product_id"]],
        "date_key": int(r["count_date"].strftime("%Y%m%d")),
        "count_batch_id": r["count_batch_id"],
        "units_on_hand": r["units_on_hand"],
    })
fact_inventory_snapshot = pd.DataFrame(fact_rows)

# ─────────────────────────────────────────────────────────────────────────────
# MESSY RAW FILES — what the learner actually receives
# ─────────────────────────────────────────────────────────────────────────────

# --- Warehouses_Raw.xlsx ---------------------------------------------------
wh_raw_rows = []
for w in WAREHOUSES:
    wh_raw_rows.append({"Warehouse_Code": w[0], "Warehouse_Name": w[1],
                         "City": w[2], "Region": w[3], "Capacity_Units": w[4]})
# duplicate WH01 with a stale (lower) capacity value from an earlier export
wh_raw_rows.append({"Warehouse_Code": "WH01", "Warehouse_Name": "Mumbai Distribution Center",
                     "City": "Mumbai", "Region": "West", "Capacity_Units": 47000})
# inconsistent casing on two rows (mutate copies already appended, so re-find them)
for row in wh_raw_rows:
    if row["Warehouse_Code"] == "WH03":
        row["City"] = "bengaluru"
    if row["Warehouse_Code"] == "WH04":
        row["City"] = "HYDERABAD"
# blank region on WH05
for row in wh_raw_rows:
    if row["Warehouse_Code"] == "WH05":
        row["Region"] = None
random.shuffle(wh_raw_rows)
warehouses_raw_df = pd.DataFrame(wh_raw_rows)

# --- Products_Raw.xlsx ------------------------------------------------------
prod_raw_rows = [{"Product_ID": p[0], "Product_Name": p[1], "Category": p[2],
                   "Sub_Category": p[3], "Unit_Cost": p[4]} for p in products]
# 3 exact duplicate rows (simulate a re-run export appending instead of overwriting)
for pcode in ["PRD-1001", "PRD-1010", "PRD-1022"]:
    match = next(r for r in prod_raw_rows if r["Product_ID"] == pcode)
    prod_raw_rows.append(dict(match))
# 2 near-duplicates: same product, category casing drifted between two exports
for pcode in ["PRD-1004", "PRD-1015"]:
    match = next(r for r in prod_raw_rows if r["Product_ID"] == pcode)
    near_dupe = dict(match)
    near_dupe["Category"] = near_dupe["Category"].lower()
    prod_raw_rows.append(near_dupe)
random.shuffle(prod_raw_rows)
products_raw_df = pd.DataFrame(prod_raw_rows)

# --- NightlyCounts_Raw.xlsx --------------------------------------------------
DATE_FORMATS = ["iso", "us_slash", "long_text"]


def format_date_messy(d, fmt):
    if fmt == "iso":
        return d  # real Excel date type
    elif fmt == "us_slash":
        return d.strftime("%m/%d/%Y")  # text
    elif fmt == "long_text":
        return f"{d.strftime('%B')} {d.day}, {d.strftime('%Y')}" if hasattr(d, "strftime") else str(d)
    return d


counts_raw_rows = []
for i, r in enumerate(true_fact_rows):
    fmt_choice = random.choices(DATE_FORMATS, weights=[0.55, 0.25, 0.20])[0]
    counts_raw_rows.append({
        "Count_Batch_ID": r["count_batch_id"],
        "Warehouse_Code": r["warehouse_id"],
        "Product_ID": r["product_id"],
        "Count_Date": format_date_messy(r["count_date"], fmt_choice),
        "Units_On_Hand": r["units_on_hand"],
    })

# ~20 duplicate rows: pick 20 random existing rows and repeat them verbatim
dupe_sample_idx = random.sample(range(len(counts_raw_rows)), 20)
for idx in dupe_sample_idx:
    counts_raw_rows.append(dict(counts_raw_rows[idx]))

# ~12 rows with blank Warehouse_Code (unresolvable -> must be excluded/flagged)
blank_wh_idx = random.sample(range(len(true_fact_rows)), 12)
for idx in blank_wh_idx:
    row = dict(counts_raw_rows[idx])
    row["Warehouse_Code"] = None
    counts_raw_rows.append(row)

# ~10 rows with a negative / invalid Units_On_Hand (data-entry error)
bad_units_idx = random.sample(range(len(true_fact_rows)), 10)
for idx in bad_units_idx:
    row = dict(counts_raw_rows[idx])
    row["Units_On_Hand"] = -1 * abs(random.randint(1, 50))
    counts_raw_rows.append(row)

random.shuffle(counts_raw_rows)
counts_raw_df = pd.DataFrame(counts_raw_rows)

TRUE_ROW_COUNT = len(true_fact_df)
DUPLICATE_ROWS_ADDED = len(dupe_sample_idx)
BLANK_WAREHOUSE_ROWS_ADDED = len(blank_wh_idx)
NEGATIVE_UNITS_ROWS_ADDED = len(bad_units_idx)
RAW_TOTAL_ROWS = len(counts_raw_df)

print(f"True clean fact grain           : {TRUE_ROW_COUNT} rows")
print(f"Duplicate rows injected         : {DUPLICATE_ROWS_ADDED}")
print(f"Blank-warehouse rows injected   : {BLANK_WAREHOUSE_ROWS_ADDED}")
print(f"Negative-units rows injected    : {NEGATIVE_UNITS_ROWS_ADDED}")
print(f"NightlyCounts_Raw.xlsx row count: {RAW_TOTAL_ROWS}")
print(f"Warehouses_Raw.xlsx row count   : {len(warehouses_raw_df)} (5 true + 1 duplicate)")
print(f"Products_Raw.xlsx row count     : {len(products_raw_df)} ({NUM_PRODUCTS} true + 3 exact dupes + 2 near-dupes)")


# ─────────────────────────────────────────────────────────────────────────────
# WRITE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def style_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 3, 12), 40)


def write_df_sheet(ws, df):
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))
    style_header(ws)


def save_single_sheet(df, filename, sheet_name):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    write_df_sheet(ws, df)
    wb.save(f"{OUT_DIR}/{filename}")
    print(f"wrote {filename}")


# ─────────────────────────────────────────────────────────────────────────────
# 1) RAW INPUT FILES (learner-facing)
# ─────────────────────────────────────────────────────────────────────────────

save_single_sheet(warehouses_raw_df, "Warehouses_Raw.xlsx", "Warehouses")
save_single_sheet(products_raw_df, "Products_Raw.xlsx", "Products")
save_single_sheet(counts_raw_df, "NightlyCounts_Raw.xlsx", "NightlyCounts")

# ─────────────────────────────────────────────────────────────────────────────
# 2) STARTER WORKBOOK (blank shells, headers only, for learners to fill in)
# ─────────────────────────────────────────────────────────────────────────────

starter_wb = Workbook()
starter_wb.remove(starter_wb.active)

for sheet_name, cols in [
    ("dim_warehouse", ["warehouse_sk", "warehouse_id", "warehouse_name", "city", "region", "capacity_units"]),
    ("dim_product", ["product_sk", "product_id", "product_name", "category", "sub_category", "unit_cost"]),
    ("dim_date", ["date_key", "date", "day_of_week", "month", "quarter", "year"]),
    ("fact_inventory_snapshot", ["warehouse_sk", "product_sk", "date_key", "count_batch_id", "units_on_hand"]),
]:
    ws = starter_wb.create_sheet(sheet_name)
    ws.append(cols)
    style_header(ws)

instructions_ws = starter_wb.create_sheet("READ ME FIRST", 0)
instructions_ws.append(["This workbook is where YOUR star schema goes."])
instructions_ws.append([""])
instructions_ws.append(["Tabs in this workbook: dim_warehouse, dim_product, dim_date, fact_inventory_snapshot"])
instructions_ws.append(["Each tab already has the exact column headers your finished table needs."])
instructions_ws.append(["Build each one using the raw files (Warehouses_Raw.xlsx, Products_Raw.xlsx,"])
instructions_ws.append(["NightlyCounts_Raw.xlsx) and the instructions document. Do not rename any tab."])
instructions_ws.column_dimensions["A"].width = 95
instructions_ws["A1"].font = Font(bold=True, size=13, color="1D4ED8")

starter_wb.save(f"{OUT_DIR}/Starter_Workbook.xlsx")
print("wrote Starter_Workbook.xlsx")

# ─────────────────────────────────────────────────────────────────────────────
# 3) ANSWER KEY (fully solved, plus a DQ summary sheet)
# ─────────────────────────────────────────────────────────────────────────────

answer_wb = Workbook()
answer_wb.remove(answer_wb.active)

dq_ws = answer_wb.create_sheet("DQ_Summary", 0)
dq_rows = [
    ["Data Quality Decisions — What Was Found and What Was Done", ""],
    ["", ""],
    ["Source file", "Issue found & resolution"],
    ["Warehouses_Raw.xlsx", "1 duplicate row for WH01 with a stale capacity_units value (47,000 vs the correct 50,000). "
                            "Kept the row with the higher/more recent capacity value; dropped the stale duplicate."],
    ["Warehouses_Raw.xlsx", "City casing inconsistent (\"bengaluru\", \"HYDERABAD\"). Standardized to proper case."],
    ["Warehouses_Raw.xlsx", "1 blank Region value (WH05). Filled in from known warehouse location (Pune -> West), "
                            "since region is derivable and not ambiguous here."],
    ["Products_Raw.xlsx", f"3 exact duplicate rows (re-export artifact). Removed with Remove Duplicates."],
    ["Products_Raw.xlsx", "2 near-duplicate rows with inconsistent Category casing (e.g. \"electronics\" vs \"Electronics\"). "
                          "Treated as the same product; kept the properly-cased version, dropped the other."],
    ["NightlyCounts_Raw.xlsx", f"{DUPLICATE_ROWS_ADDED} exact duplicate count rows (same warehouse+product+date+units "
                               "repeated). Removed with Remove Duplicates before building the fact table."],
    ["NightlyCounts_Raw.xlsx", f"{BLANK_WAREHOUSE_ROWS_ADDED} rows had a blank Warehouse_Code with no way to recover which "
                               "warehouse they belonged to. These cannot be safely joined to dim_warehouse, so they were "
                               "EXCLUDED (quarantined) from fact_inventory_snapshot rather than guessed."],
    ["NightlyCounts_Raw.xlsx", f"{NEGATIVE_UNITS_ROWS_ADDED} rows had a negative Units_On_Hand value (a physically impossible "
                               "count -- a data-entry error). EXCLUDED from fact_inventory_snapshot."],
    ["NightlyCounts_Raw.xlsx", "Count_Date arrived in 3 different formats across rows (ISO, US slash MM/DD/YYYY, and long "
                               "text like \"June 1, 2026\") -- a realistic symptom of the raw export coming from more than "
                               "one system over time. All values were normalized to a single real Excel date before building "
                               "date_key."],
    ["", ""],
    ["Result", f"True clean grain is {TRUE_ROW_COUNT} rows (5 warehouses x {NUM_PRODUCTS} products x {NUM_NIGHTS} nights). "
               f"After removing duplicates ({DUPLICATE_ROWS_ADDED}) and excluding unresolvable/invalid rows "
               f"({BLANK_WAREHOUSE_ROWS_ADDED} blank warehouse + {NEGATIVE_UNITS_ROWS_ADDED} negative units), "
               f"fact_inventory_snapshot should land back at exactly {TRUE_ROW_COUNT} rows. "
               "If your row count differs, you likely double-counted a duplicate or kept an invalid row."],
]
for row in dq_rows:
    dq_ws.append(row)
dq_ws["A1"].font = Font(bold=True, size=13, color="1D4ED8")
dq_ws.column_dimensions["A"].width = 22
dq_ws.column_dimensions["B"].width = 110
for r in dq_ws.iter_rows(min_row=3):
    for c in r:
        c.alignment = Alignment(wrap_text=True, vertical="top")

for sheet_name, df in [
    ("dim_warehouse", dim_warehouse),
    ("dim_product", dim_product),
    ("dim_date", dim_date),
    ("fact_inventory_snapshot", fact_inventory_snapshot),
]:
    ws = answer_wb.create_sheet(sheet_name)
    write_df_sheet(ws, df)

answer_wb.save(f"{OUT_DIR}/Answer_Key_Star_Schema.xlsx")
print("wrote Answer_Key_Star_Schema.xlsx")

print("\nDONE. All artifacts written to:", OUT_DIR)
