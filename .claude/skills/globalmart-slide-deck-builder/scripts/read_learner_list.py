"""
read_learner_list.py — load a cohort's learner roster and deterministically
rotate a real name into a worked-example slide.

Why this exists: the slide deck's tone uses a running named-learner example
("Sneha needs to find last month's top-selling category") the same way the
original reference deck's "Rahul" example did. Using the SAME name every
single session gets stale fast, but picking a name at random makes a deck
non-reproducible (running this skill twice for the same session should not
silently change which learner appears on it). pick_learner() solves both:
deterministic per (day, session_seed), varied across different sessions.

Default cohort is Hyderabad 2026 (the operative batch as of 2026-08-12).
Pass --learner-list to point at a different cohort's roster, e.g. the
original Chennai 2026 one.

Usage:
    python read_learner_list.py                       # print full roster
    python read_learner_list.py --pick 7 ILT2          # pick a learner for Day 7, ILT2
    python read_learner_list.py --learner-list "..\\..\\..\\..\\calendars\\tred-alch-adv-dbx-chennai-2026-learner list (duplicate, identical to the other copy).xlsx" --pick 7 ILT2
"""

import argparse
import hashlib
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# 2026-07-19: Chennai_batch was merged into this repo; the learner list now
# lives in calendars/ (an exact-duplicate copy also sits alongside it,
# labeled "(duplicate, identical to the other copy)" -- either is fine, this
# one is the original).
#
# 2026-08-12: default cohort switched to Hyderabad 2026. NOTE: the Chennai
# roster file this constant used to point at (and the absolute
# C:\Yvirinchy\tred-alch-adv-dbx\ root it lived under) has since been deleted
# from the repo -- only the "(duplicate, identical to the other copy)" file
# remains for Chennai. Path is now relative to this script so it survives
# repo moves (the old absolute path silently pointed at a dead root).
LEARNER_LIST_PATH = str(
    Path(__file__).resolve().parents[4] / "calendars" / "tred-alch-adv-dbx-hyderabad-2026-learner-list.xlsx"
)
SHEET_NAME = "Sheet1"


def load_learners(path: str = LEARNER_LIST_PATH, sheet: str = SHEET_NAME):
    """Return a list of (name, email) tuples, in row order, header excluded."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    name_idx = header.index("name") if "name" in header else 0
    email_idx = header.index("email") if "email" in header else 1
    learners = []
    for row in rows[1:]:
        if row is None or len(row) <= max(name_idx, email_idx):
            continue
        name, email = row[name_idx], row[email_idx]
        if name is None:
            continue
        learners.append((str(name).strip(), str(email).strip() if email else ""))
    return learners


def pick_learner(day, session_seed, path: str = LEARNER_LIST_PATH, sheet: str = SHEET_NAME):
    """
    Deterministically pick one (name, email) tuple for a given (day, session_seed).

    day: int or str, e.g. 7 or "Day7"
    session_seed: any string identifying the session, e.g. "ILT2",
                  "HOL1_Star_Schema_Design", or the notebook's basename.

    Same (day, session_seed) always returns the same learner -- rerunning this
    skill for the same session is reproducible. Different sessions land on a
    different index almost all the time (25 learners, so collisions across
    unrelated sessions are expected and fine -- this is rotation for variety,
    not a uniqueness guarantee).
    """
    learners = load_learners(path, sheet)
    if not learners:
        raise ValueError(f"No learners loaded from {path} sheet {sheet!r}")
    key = f"Day{day}|{session_seed}".encode("utf-8")
    digest = hashlib.sha256(key).hexdigest()
    index = int(digest, 16) % len(learners)
    return learners[index]


def _self_test():
    """Quick sanity check: same call is stable, different calls vary."""
    cases = [
        (1, "ILT1"), (1, "HOL"), (3, "HOL1"), (6, "HOL1"),
        (7, "HOL2"), (9, "ILT1"), (13, "ILT"),
    ]
    print(f"{'day':<6}{'session':<20}{'name':<24}{'email'}")
    results = {}
    for day, session in cases:
        learner = pick_learner(day, session)
        results[(day, session)] = learner
        print(f"{day:<6}{session:<20}{learner[0]:<24}{learner[1]}")

    # Determinism check: calling again with the same key returns the same result.
    for (day, session), learner in results.items():
        again = pick_learner(day, session)
        assert again == learner, f"Non-deterministic result for Day{day}/{session}: {learner} vs {again}"
    print("\nDeterminism check passed: repeated calls with the same (day, session) match.")

    names = {v[0] for v in results.values()}
    print(f"Distinct names across {len(results)} test cases: {len(names)}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--pick", nargs=2, metavar=("DAY", "SESSION_SEED"),
                         help="Pick a learner for the given day and session seed.")
    parser.add_argument("--self-test", action="store_true",
                         help="Run the built-in determinism/variety self-test.")
    parser.add_argument("--learner-list", type=str, default=LEARNER_LIST_PATH,
                         help="Override path to the roster .xlsx (e.g. the Chennai roster path)")
    args = parser.parse_args()

    if args.self_test:
        _self_test()
    elif args.pick:
        day, session_seed = args.pick
        name, email = pick_learner(day, session_seed, path=args.learner_list)
        print(f"{name} <{email}>")
    else:
        for name, email in load_learners(path=args.learner_list):
            print(f"{name}\t{email}")
