#!/usr/bin/env python3
"""
read_calendar.py -- read one calendar day out of a GlobalMart cohort calendar
and compute the {order}/{type}/{x} naming fields for its ILT/Hands-on rows.

Default cohort is Hyderabad 2026 (the operative batch as of 2026-08-12).
Pass --calendar to point at a different cohort's calendar file, e.g. the
original Chennai 2026 one (calendars/tred-alch-adv-dbx-chennai-2026-calendar.xlsx).

IMPORTANT: "Day Number" is NOT on the same scale across cohorts. Chennai's
calendar numbers the Databricks track 1-13 (matching the Databricks/Day{N}/
folders on disk). Hyderabad's calendar numbers the SAME Databricks track
9-22 (continuous with its SQL & Python track, which occupies 1-8) -- e.g.
Hyderabad Day 9 = Chennai Day 1 (GlobalMart Problem Statement/Medallion),
Hyderabad Day 21 = Chennai Day 12 (UC Governance/Genie), etc, with one extra
Hyderabad-only day (16: Z-order vs Liquid Clustering + "Design Your Own
Architecture") that has no Chennai equivalent. This script does NOT remap
Hyderabad day numbers onto the existing Day1-13 folders -- if you're using
this script's output to name/save a file, you must decide by hand which
on-disk Day{N} folder a given Hyderabad --day maps to (or whether it needs
a new folder), the same way you'd reconcile order numbers per the CAVEAT
below.

Why this exists: the calendar workbook only fills in Week Number / Day Number /
Date / Day-name on the FIRST row of each session block (everything else is
blank, meant to be read visually in Excel with merged-looking cells). This
script forward-fills those four columns so each row knows which Day{N} it
belongs to, then filters to the rows that actually turn into a notebook
(Activity Type == "ILT" or "Hands-on" -- Lunch Break / Assessment Prep /
Assessment (Project: NN) / Tredence Internal Sessions rows are dropped).

Two real-world wrinkles this script handles that a naive "just count the rows"
approach gets wrong -- both confirmed by reading the actual workbook and
cross-checking against the real files already built in Day1-Day12:

1. A single session sometimes spans more than one calendar time-block, and
   shows up as two (or more) consecutive rows with IDENTICAL Module text.
   Day 9's "Handling incremental data - Bronze & Silver" hands-on is exactly
   this -- two back-to-back rows, same text, but only ONE real file
   (Day9_5_HOL2...) was built. This script merges consecutive rows with the
   same normalized Module text into a single logical session before counting,
   printing "[merged N rows]" so the merge is visible and auditable rather
   than silently changing the count.

2. The calendar can have a genuinely malformed row: Day 3's row for "Connect
   to Public API + GraphDB..." appears once tagged Hands-on, then AGAIN on the
   very next row tagged ILT with the identical Module text. That's a data-
   entry duplicate, not two different sessions. The same identical-text merge
   rule above collapses it into one HOL row, using the FIRST occurrence's
   Activity Type (Hands-on) -- which matches what was actually built
   (Day3_2_HOL1), not a phantom second ILT.

IMPORTANT CAVEAT this script cannot solve on its own: {order} computed here is
relative to calendar rows ONLY. Real Day folders also contain DEMO/REF files
(cert-prep, instructor-only demos) that are NOT in the calendar at all, and
those consume order slots too -- e.g. Day 3's real order is
ILT1=1, HOL1=2, ILT2=3, DEMO=4, HOL2=5, REF=6, but this script (seeing only
4 calendar-derived rows) would compute HOL2 as order 4, not 5. Always
reconcile this script's output against the actual Day{N} folder (via Glob)
before trusting a number for a day that might have a DEMO/REF file -- this
script gives you the right RELATIVE order among calendar-derived sessions,
not necessarily the final absolute filename number.

Usage:
    python read_calendar.py --day 9
    python read_calendar.py --day 13 --all-rows

Dependencies: openpyxl only (pip install openpyxl if missing).
"""
import argparse
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: this script needs openpyxl. Install it with: pip install openpyxl",
          file=sys.stderr)
    sys.exit(1)

# Reconfigure stdout to UTF-8 where possible -- calendar Module text and Day7-
# style em dashes / arrows can otherwise crash a cp1252 Windows console.
try:
    sys.stdout.reconfigure(encoding="utf-8")
except AttributeError:
    pass

# 2026-07-19: calendar moved from repo root into calendars/. The Chennai
# calendar's session titles don't match the real built notebooks verbatim
# (e.g. this file's own session titles use "Problem Statement & GlobalMart
# Architecture", which matches neither calendar's Day-1 title exactly).
# Treat this script's {order}/{x} output as a strong hint, not gospel --
# always cross-check with Glob against the real Day{N} folder before
# finalizing a filename, same as the existing REMINDER at the bottom of
# this script already says.
#
# 2026-08-12: default cohort switched to Hyderabad 2026. The Chennai
# alt-version calendar file this used to also point at has since been
# deleted from the repo (git status showed it removed) -- pass --calendar
# explicitly if you need to read a specific cohort's file instead of the
# default.
CALENDAR_PATH = Path(__file__).resolve().parents[4] / "calendars" / "tred-alch-adv-dbx-hyderabad-2026-calendar.xlsx"
CHENNAI_CALENDAR_PATH = Path(__file__).resolve().parents[4] / "calendars" / "tred-alch-adv-dbx-chennai-2026-calendar.xlsx"
SHEET_NAME = "Databricks (DE)"

# Activity Type values that actually become a notebook file. Everything else
# (Lunch Break, Assessment Prep, Assessment (Project: NN), Tredence Internal
# Sessions, and blank/weekend placeholder rows) is out of scope for this skill.
SESSION_TYPES = {"ilt": "ILT", "hands-on": "Hands-on"}


def load_rows(path):
    """Load the sheet and forward-fill Week Number/Day Number/Date/Day columns
    (A-D). Columns E-H (Time/Activity Type/Module/Trainer) are already
    populated per-row in the source and are NOT forward-filled."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)
    ws = wb[SHEET_NAME]

    rows = []
    last = {"week": None, "day_num": None, "date": None, "day_name": None}
    for r in ws.iter_rows(min_row=2, values_only=True):  # skip header row
        week, day_num, date, day_name, time_, activity, module, trainer = (list(r) + [None] * 8)[:8]

        # Weekend placeholder rows (raw Day == 'Saturday'/'Sunday', everything
        # else blank) are NOT part of any Day{N} -- they're just visual spacers
        # in the sheet. Forward-filling would wrongly attach them to the
        # preceding Friday's Day Number. Drop them outright instead.
        if day_name in ("Saturday", "Sunday") and day_num is None and week is None:
            continue

        if week is not None:
            last["week"] = week
        if day_num is not None:
            last["day_num"] = day_num
        if date is not None:
            last["date"] = date
        if day_name is not None:
            last["day_name"] = day_name
        rows.append({
            "week": last["week"],
            "day_num": last["day_num"],
            "date": last["date"],
            "day_name": last["day_name"],
            "time": time_,
            "activity": activity,
            "module": module,
            "trainer": trainer,
        })
    return rows


def normalize_module(text):
    if text is None:
        return ""
    return " ".join(str(text).split()).strip().lower()


def filter_session_rows(rows):
    """Keep only ILT/Hands-on rows, dropping Lunch/Assessment/Internal/blank
    rows. Do this BEFORE merging duplicates -- see merge_consecutive_duplicates
    for why order matters here."""
    return [r for r in rows if str(r["activity"] or "").strip().lower() in SESSION_TYPES]


def merge_consecutive_duplicates(rows):
    """Collapse consecutive rows with identical normalized Module text into a
    single logical session (see module docstring, wrinkle #1/#2). Keeps the
    FIRST row's activity type and Module text; combines the time range as
    'start of first - end of last'.

    MUST be called on an already-filtered (ILT/Hands-on only) list, not the
    raw per-day row list. Reason: a split session's two halves are not always
    literally back-to-back rows in the raw sheet -- Day 10's "Incremental Gold
    Refresh - MERGE-based fact table updates" Hands-on appears once before
    Lunch and again after it, with the Lunch row sitting physically between
    them. Filtering Lunch out first makes the two halves adjacent, so this
    function's plain "same text as the row right before it" rule can still
    catch it. Merging on the raw (unfiltered) list would miss this pair and
    silently over-count Day 10 by one phantom HOL -- confirmed by hand against
    the real Day10\\ folder (5 real files: ILT1, HOL1, HOL2, ILT2, ILT3; a
    pre-filter merge produces the correct HOL1/HOL2 pair, a post-filter-only
    check without this ordering does not)."""
    merged = []
    i = 0
    while i < len(rows):
        group = [rows[i]]
        j = i + 1
        norm = normalize_module(rows[i]["module"])
        while j < len(rows) and normalize_module(rows[j]["module"]) == norm and norm != "":
            group.append(rows[j])
            j += 1
        if len(group) > 1:
            start_time = group[0]["time"].split("-")[0].strip() if group[0]["time"] else "?"
            end_time = group[-1]["time"].split("-")[-1].strip() if group[-1]["time"] else "?"
            merged_row = dict(group[0])
            merged_row["time"] = f"{start_time} - {end_time}"
            merged_row["_merged_count"] = len(group)
            merged.append(merged_row)
        else:
            merged.append(group[0])
        i = j
    return merged


def rows_for_day(rows, day_number):
    return [r for r in rows if r["day_num"] == float(day_number)]


def compute_naming(session_rows):
    """Given a day's rows that are ALREADY filtered to ILT/Hands-on AND
    already merged (see filter_session_rows / merge_consecutive_duplicates,
    called in that order), compute order/type/x. order = 1-based position
    among THIS day's ILT/Hands-on rows only. x = 1-based position within that
    type, dropped when only one session of that type exists this day."""
    type_counts = {}
    for r in session_rows:
        key = SESSION_TYPES[str(r["activity"]).strip().lower()]
        type_counts[key] = type_counts.get(key, 0) + 1

    type_seen = {}
    named = []
    for order, r in enumerate(session_rows, start=1):
        type_key = SESSION_TYPES[str(r["activity"]).strip().lower()]
        short_type = "ILT" if type_key == "ILT" else "HOL"
        type_seen[type_key] = type_seen.get(type_key, 0) + 1
        x = type_seen[type_key]
        drop_x = type_counts[type_key] == 1
        label = short_type if drop_x else f"{short_type}{x}"
        named.append({**r, "order": order, "type": short_type, "x": None if drop_x else x, "label": label})
    return named


def main():
    ap = argparse.ArgumentParser(description="Read one Day's calendar rows and compute naming fields.")
    ap.add_argument("--day", type=int, required=True,
                     help="Day Number as it appears in the calendar's 'Day Number' column. "
                          "Hyderabad (default calendar): 9-22. Chennai (--calendar override): 1-13, "
                          "matches Databricks/Day{N} folder -- see module docstring for the mapping "
                          "between the two.")
    ap.add_argument("--all-rows", action="store_true", help="Show every row for that day, not just ILT/Hands-on")
    ap.add_argument("--calendar", type=str, default=None,
                     help="Override path to the calendar .xlsx (e.g. the Chennai calendar path)")
    args = ap.parse_args()

    path = Path(args.calendar) if args.calendar else CALENDAR_PATH
    if not path.exists():
        print(f"ERROR: calendar file not found at {path}", file=sys.stderr)
        sys.exit(1)

    all_rows = load_rows(path)
    day_rows_raw = rows_for_day(all_rows, args.day)
    if not day_rows_raw:
        print(f"No rows found for Day {args.day}. Valid Day Numbers depend on the calendar in use: "
              f"Hyderabad (default) runs 9-22, Chennai runs 1-13 (weekends have none either way).")
        sys.exit(1)

    date_val = day_rows_raw[0]["date"]
    date_str = date_val.strftime("%Y-%m-%d") if hasattr(date_val, "strftime") else str(date_val)
    print(f"Day {args.day} -- {day_rows_raw[0]['day_name']} {date_str} (Week {day_rows_raw[0]['week']})")
    print("=" * 70)

    if args.all_rows:
        for r in day_rows_raw:
            print(f"  {r['time']:<22} {str(r['activity']):<28} {r['module']}")
        print()

    # Filter to ILT/Hands-on FIRST, then merge duplicates -- in that order.
    # Merging before filtering would miss splits like Day 10's, where a Lunch
    # row sits physically between the two halves of one Hands-on session.
    # See merge_consecutive_duplicates' docstring for the full explanation.
    session_rows = merge_consecutive_duplicates(filter_session_rows(day_rows_raw))

    named = compute_naming(session_rows)
    if not named:
        print("No ILT/Hands-on rows this day (e.g. Day 8/13 are assessment-only).")
        print("Nothing for globalmart-notebook-builder to build here.")
        return

    print(f"\n{'order':<7}{'type':<6}{'x':<5}{'label':<8}{'time':<22}module")
    print("-" * 100)
    for n in named:
        merged_note = f" [merged {n['_merged_count']} rows]" if n.get("_merged_count") else ""
        x_display = "" if n["x"] is None else n["x"]
        print(f"{n['order']:<7}{n['type']:<6}{str(x_display):<5}{n['label']:<8}{n['time']:<22}{n['module']}{merged_note}")

    print()
    print("REMINDER: {order} above is relative to calendar rows only. Two things can make")
    print("the real on-disk number different from what's printed here:")
    print("  1. A DEMO or REF file already in this Day's folder (not in the calendar at")
    print("     all) shifts every order number after it -- see Day 3 in naming-convention.md.")
    print("  2. A session split across a non-adjacent gap (e.g. a Hands-on split by Lunch,")
    print("     or spilling into the next calendar day's first row) can still throw off the")
    print("     count in edge cases this script can't fully see -- see Day 10 and Day 6 in")
    print("     naming-convention.md. Always cross-check with Glob against the real")
    print("     Day{N}\\ folder before finalizing a new filename.")


if __name__ == "__main__":
    main()
